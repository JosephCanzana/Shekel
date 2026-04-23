import io
import json
from collections import Counter, defaultdict
from datetime import datetime,  timedelta , date as _date
import csv
import io
from datetime import datetime

from flask import (
    Blueprint, render_template, request,
    jsonify, url_for, send_file, redirect, flash, Response
)
from flask_login import login_required, current_user
from sqlalchemy import func

from app.utils.decorator import role_required
from app.utils.index_helpers import (
    get_admin_stats, get_time_of_day, get_low_stock_items, get_defects,
)
from app.utils.helpers import to_pht, _pht_fix

from app.models.stock_adjustment_request import StockAdjustmentRequest
from app.models.defect_detail import DefectDetail
from app.models.defect     import Defect
from app.models.product    import Product
from app.models.user       import User
from app.models.inventory  import Inventory
from app.models.stock_in   import StockIn
from app.models.audit_log  import AuditLog
from app.models.category import Category
from app.extensions import db
from app.utils.navbar_notifications import get_navbar_notifications

admin_bp = Blueprint("admin", __name__, url_prefix='/admin')


def _serialize_exchange_items(detail):
    return [
    {
    "product_name":     ei.product.product_name.capitalize() if ei.product else ei.product_id,
    "product_id":       ei.product_id,
    "quantity":         ei.quantity,
    "price_at_exchange": float(ei.price_at_exchange),
    "no_money_exchange": ei.no_money_exchange,
    "override_used":     ei.override_used,
    }
    for ei in (detail.exchange_items or [])
    ]

# ── Dashboard ─────────────────────────────────────────────────────────────────
@admin_bp.route("/")
@login_required
@role_required("superadmin", "admin")
def dashboard():
    
    date_from, date_to, date_from_str, date_to_str = _parse_date_range()
    active_tab = request.args.get("tab",  "sales")
    page       = request.args.get("page", 1, type=int)
 
    data       = _build_report_data(date_from, date_to)
    chart_json = json.dumps(_build_chart_data(data))
    stats      = get_admin_stats()
 
    pending = (
        StockAdjustmentRequest.query
        .filter_by(status="pending")
        .order_by(StockAdjustmentRequest.submitted_at.asc())
        .limit(3)
        .all()
    )
 
    # ── Detail rows (paginated table per tab) ─────────────────────────────────
    # Sales tab: paginate the *daily* breakdown (not hourly).
    # The template's breakdown table should iterate detail.rows;
    # for single-day it falls through to data.sales.hourly in the Jinja block.
    if active_tab == "inventory":
        detail = _paginate_list(data["inventory"]["rows"], page)
    elif active_tab == "stock":
        detail = _paginate_list(data["stock"]["rows"], page)
    elif active_tab == "defects":
        detail = _paginate_list(data["defects"]["rows"], page)
    else:  # sales
        is_single_day = (date_from_str == date_to_str)
        if is_single_day and data["sales"]["hourly"]:
            detail = _paginate_list(data["sales"]["hourly"], page)
        else:
            detail = _paginate_list(list(data["sales"]["daily"]), page)
 
    return render_template(
        "admin/reports.html",
        time_of_day              = get_time_of_day(),
        stats                    = stats,
        data                     = data,
        detail                   = detail,
        chart_json               = chart_json,
        active_tab               = active_tab,
        date_from                = date_from_str,
        date_to                  = date_to_str,
        low_stock_items          = get_low_stock_items(),
        defects                  = get_defects(),
        pending_requests_preview = [r.to_dict() for r in pending],
    )

# ── Approval page ─────────────────────────────────────────────────────────────
@admin_bp.route("/requests")
@login_required
@role_required("superadmin", "admin")
def requests_page():
    pending = StockAdjustmentRequest.query\
        .filter_by(status="pending")\
        .order_by(StockAdjustmentRequest.submitted_at.desc())\
        .all()
    history = StockAdjustmentRequest.query\
        .filter(StockAdjustmentRequest.status != "pending")\
        .order_by(StockAdjustmentRequest.reviewed_at.desc())\
        .limit(30).all()

    # Group submitted defect details by defect header
    defect_rows = (
        db.session.query(DefectDetail, Defect, Product, User)
        .join(Defect,   Defect.defect_id     == DefectDetail.defect_id)
        .join(Product,  Product.product_id   == DefectDetail.product_id)
        .join(User,     User.user_id         == Defect.user_id)
        .filter(DefectDetail.status      == "submitted")
        .filter(DefectDetail.is_archived == False)
        .order_by(Defect.defect_datetime.asc(), DefectDetail.defect_detail_id.asc())
        .all()
    )

    # Group by defect_id so one card = one log session
    from collections import OrderedDict
    defect_groups = OrderedDict()
    for detail, defect, product, user in defect_rows:
        did = defect.defect_id
        if did not in defect_groups:
            defect_groups[did] = {
                "defect_id":  defect.defect_id,
                "datetime":   to_pht(defect.defect_datetime).strftime("%b %d, %Y %I:%M %p"),
                "logged_by":  f"{user.first_name} {user.last_name}".strip().title(),
                "logged_role": user.role,
                "items": [],
            }
        defect_groups[did]["items"].append({
            "detail_id":             detail.defect_detail_id,
            "product_id":            product.product_id,
            "product_name":          product.product_name.capitalize(),
            "quantity":              detail.quantity,
            "price_at_defect":       float(detail.price_at_defect),
            "subtotal_amount":       float(detail.subtotal_amount),
            "origin":                detail.origin,
            "origin_label":          "Customer" if detail.origin == "customer" else "In-Store",
            "reason":                detail.reason,
            "reason_label":          detail.reason.replace("_", " ").title(),
            "customer_compensation": detail.customer_compensation.replace("_", " ").title(),
            "customer_compensation_raw": detail.customer_compensation,
            "transaction_id":        f"TXN-{detail.transaction_id:05d}" if detail.transaction_id else None,
            "exchange_items":        _serialize_exchange_items(detail),
            "approve_url":  url_for("defects.approve",        detail_id=detail.defect_detail_id),
            "reject_url":   url_for("defects.reject",         detail_id=detail.defect_detail_id),
            "delete_url":   url_for("defects.archive_detail", detail_id=detail.defect_detail_id),
        })

    defect_data = list(defect_groups.values())

    proposal_rows = (
        db.session.query(DefectDetail, Defect, Product, User)
        .join(Defect,   Defect.defect_id     == DefectDetail.defect_id)
        .join(Product,  Product.product_id   == DefectDetail.product_id)
        .join(User,     User.user_id         == Defect.user_id)
        .filter(DefectDetail.status                        == "active")
        .filter(DefectDetail.supplier_compensation         == "pending")
        .filter(DefectDetail.proposed_supplier_compensation != None)
        .filter(DefectDetail.is_archived                   == False)
        .order_by(Defect.defect_datetime.asc())
        .all()
    )

    proposal_data = [
        {
            "detail_id":    detail.defect_detail_id,
            "product_name": product.product_name,
            "quantity":     detail.quantity,
            "origin_label": "Customer" if detail.origin == "customer" else "In-Store",
            "reason_label": detail.reason.replace("_", " ").title(),
            "proposed":     detail.proposed_supplier_compensation.replace("_", " ").title(),
            "proposed_raw": detail.proposed_supplier_compensation,
            "logged_by":    f"{user.first_name} {user.last_name}".strip().title(),
            "datetime":     to_pht(defect.defect_datetime).strftime("%b %d, %Y %I:%M %p"),
            "approve_url":  url_for("defects.review",         detail_id=detail.defect_detail_id),
            "reject_url":   url_for("defects.clear_proposal", detail_id=detail.defect_detail_id),
        }
        for detail, defect, product, user in proposal_rows
    ]

    defect_history_rows = (
        db.session.query(DefectDetail, Defect, Product, User)
        .join(Defect,   Defect.defect_id     == DefectDetail.defect_id)
        .join(Product,  Product.product_id   == DefectDetail.product_id)
        .join(User,     User.user_id         == Defect.user_id)
        .filter(DefectDetail.status.in_(["active", "rejected"]))
        .filter(DefectDetail.is_archived == False)
        .order_by(Defect.defect_datetime.desc())
        .limit(30).all()
    )

    defect_history = [
        {
            "product_name":          product.product_name.capitalize(),
            "quantity":              detail.quantity,
            "origin_label":          "Customer" if detail.origin == "customer" else "In-Store",
            "reason_label":          detail.reason.replace("_", " ").title(),
            "customer_compensation": detail.customer_compensation.replace("_", " ").title(),
            "supplier_compensation": detail.supplier_compensation.replace("_", " ").title(),
            "status":                detail.status,
            "logged_by":             f"{user.first_name} {user.last_name}".strip().title(),
            "rejection_note":        detail.rejection_note or "",
            "datetime":              to_pht(defect.defect_datetime).strftime("%b %d, %Y %I:%M %p"),
        }
        for detail, defect, product, user in defect_history_rows
    ]

    def _history_dict(r):
        d = _pht_fix(r.to_dict())
        d['details'] = [
            {
                'product_id':         det.product_id,
                'product_name':       Product.query.get(det.product_id).product_name.capitalize(),
                'quantity_requested': det.quantity_requested,
                'quantity_approved':  det.quantity_approved,
                'status':             det.status,
                'rejection_reason':   det.rejection_reason or '',
                'note':               det.note or '',
            }
            for det in r.details
        ]
        return d

    return render_template(
        "admin/requests.html",
        pending_data   = [_pht_fix(r.to_dict()) for r in pending],
        history_data   = [_history_dict(r) for r in history],
        defect_data    = defect_data,
        proposal_data  = proposal_data,
        defect_history = defect_history,
    )


# ── Review API ────────────────────────────────────────────────────────────────
@admin_bp.route("/requests/<int:request_id>/review", methods=["POST"])
@login_required
@role_required("superadmin", "admin")
def review_request(request_id):
    req = StockAdjustmentRequest.query.get_or_404(request_id)

    if req.status != "pending":
        return jsonify({"error": "This request has already been reviewed."}), 409

    data      = request.json or {}
    decisions = data.get("decisions", [])

    if not decisions:
        return jsonify({"error": "No decisions provided."}), 400

    decision_map = {d["detail_id"]: d for d in decisions}

    last_batch = db.session.query(func.max(StockIn.batch_id)).scalar() or 0
    batch_id   = last_batch + 1

    for detail in req.details:
        decision = decision_map.get(detail.detail_id)
        if not decision:
            continue

        action  = decision.get("action")
        product = detail.product

        if action == "approve":
            raw_qty = decision.get("quantity_approved")
            try:
                qty = int(raw_qty) if raw_qty is not None else detail.quantity_requested
                if qty < 0 or (qty <= 0 and detail.request.request_type != "adjustment"):
                    raise ValueError
            except (ValueError, TypeError):
                return jsonify({"error": f"Invalid quantity for '{detail.product.product_name}'."}), 400

            detail.approve(qty)
            if detail.request.request_type == "adjustment":
                if product and product.status != "archived":
                    if product.inventory:
                        product.inventory.quantity_available = qty
                        product.inventory.last_updated        = datetime.utcnow()
            else:
                if product and product.status != "archived":
                    if product.inventory:
                        product.inventory.quantity_available += qty
                        product.inventory.last_updated        = datetime.utcnow()
                    else:
                        db.session.add(Inventory(
                            product_id         = detail.product_id,
                            quantity_available = qty,
                            quantity_defective = 0,
                            last_updated       = datetime.utcnow()
                        ))

                    db.session.add(StockIn(
                        product_id        = detail.product_id,
                        batch_id          = batch_id,
                        user_id           = req.requested_by,
                        quantity_received = qty,
                        stockin_datetime  = datetime.utcnow(),
                        notes             = detail.note or None,
                    ))

        elif action == "reject":
            reason = (decision.get("rejection_reason") or "").strip() or None
            detail.reject(reason)

    req.reviewed_by = current_user.user_id
    req.reviewed_at = datetime.utcnow()
    req.recompute_status()

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Failed to save. Please try again."}), 500

    return jsonify({
        "success":        True,
        "request_id":     req.request_id,
        "status":         req.status,
        "approved_count": req.approved_count,
        "rejected_count": req.rejected_count,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# AUDIT LOGS
# ═══════════════════════════════════════════════════════════════════════════════

@admin_bp.route("/audit_logs")
@login_required
@role_required("superadmin", "admin")
def audit_logs():
    from sqlalchemy import distinct

    PER_PAGE = 50

    # ── query params ──────────────────────────────────────────────────────────
    page         = request.args.get("page",        1,  type=int)
    raw_uid      = request.args.get("user_id",     "").strip()
    action_type  = request.args.get("action_type", "").strip()
    module       = request.args.get("module",      "").strip()
    q            = request.args.get("q",           "").strip()
    date_from    = request.args.get("date_from",   "").strip()
    date_to      = request.args.get("date_to",     "").strip()

    # ── base query ────────────────────────────────────────────────────────────
    qry = (
        AuditLog.query
        .outerjoin(User, AuditLog.user_id == User.user_id)
        .order_by(AuditLog.action_datetime.desc())
    )

    if raw_uid:
        try:
            qry = qry.filter(AuditLog.user_id == int(raw_uid))
        except ValueError:
            pass
    if action_type:
        qry = qry.filter(AuditLog.action_type == action_type)
    if module:
        qry = qry.filter(AuditLog.module == module)
    if q:
        qry = qry.filter(AuditLog.description.ilike(f"%{q}%"))
    if date_from:
        try:
            qry = qry.filter(
                AuditLog.action_datetime >= datetime.strptime(date_from, "%Y-%m-%d")
            )
        except ValueError:
            date_from = ""
    if date_to:
        try:
            qry = qry.filter(
                AuditLog.action_datetime
                <= datetime.strptime(date_to, "%Y-%m-%d").replace(
                    hour=23, minute=59, second=59
                )
            )
        except ValueError:
            date_to = ""

    logs = qry.paginate(page=page, per_page=PER_PAGE, error_out=False)

    # ── today stats ───────────────────────────────────────────────────────────
    today_start = datetime.combine(_date.today(), datetime.min.time())

    total_today = AuditLog.query.filter(
        AuditLog.action_datetime >= today_start
    ).count()

    logins_today = AuditLog.query.filter(
        AuditLog.action_datetime >= today_start,
        AuditLog.action_type == "LOGIN",
    ).count()

    unique_users_today = (
        db.session.query(db.func.count(distinct(AuditLog.user_id)))
        .filter(AuditLog.action_datetime >= today_start)
        .scalar()
        or 0
    )

    all_users = User.query.order_by(User.first_name, User.last_name).all()

    filters = dict(
        user_id     = int(raw_uid) if raw_uid and raw_uid.isdigit() else None,
        action_type = action_type,
        module      = module,
        q           = q,
        date_from   = date_from,
        date_to     = date_to,
    )

    stats = dict(
        total_today         = total_today,
        logins_today        = logins_today,
        unique_users_today  = unique_users_today,
    )

    return render_template(
        "admin/audit_logs.html",
        logs      = logs,
        filters   = filters,
        stats     = stats,
        all_users = all_users,
    )


@admin_bp.route("/audit_logs/export")
@login_required
@role_required("superadmin", "admin")
def export_audit_logs():
    """Download filtered audit log as a formatted Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    f_user      = request.args.get("user_id",     type=int)
    f_action    = request.args.get("action_type", "").strip()
    f_module    = request.args.get("module",      "").strip()
    f_search    = request.args.get("q",           "").strip()
    f_date_from = request.args.get("date_from",   "").strip()
    f_date_to   = request.args.get("date_to",     "").strip()

    q = AuditLog.query.join(User, AuditLog.user_id == User.user_id)
    if f_user:
        q = q.filter(AuditLog.user_id == f_user)
    if f_action:
        q = q.filter(AuditLog.action_type == f_action)
    if f_module:
        q = q.filter(AuditLog.module == f_module)
    if f_search:
        q = q.filter(AuditLog.description.ilike(f"%{f_search}%"))
    if f_date_from:
        try:
            q = q.filter(AuditLog.action_datetime >= datetime.strptime(f_date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if f_date_to:
        try:
            dt_to = datetime.strptime(f_date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            q = q.filter(AuditLog.action_datetime <= dt_to)
        except ValueError:
            pass

    logs = q.order_by(AuditLog.action_datetime.desc()).all()

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Audit Log"

    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    thin     = Border(bottom=Side(style="thin", color="E2E8F0"))

    headers    = ["#", "Date & Time (PHT)", "User", "Role", "Action", "Module", "Description", "Reference"]
    col_widths = [6,   22,                  24,     12,     10,       12,       60,             18]

    from openpyxl.utils import get_column_letter
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, log in enumerate(logs, 2):
        pht_dt = to_pht(log.action_datetime).strftime("%b %d, %Y %I:%M %p") if log.action_datetime else ""
        u      = log.user
        name   = f"{u.first_name} {u.last_name}".strip().title() if u else "—"
        role   = u.role.title() if u else "—"
        ref    = f"{log.reference_table} #{log.reference_id}" if log.reference_id else "—"

        row_data = [ri - 1, pht_dt, name, role, log.action_type, log.module, log.description, ref]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = thin
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 7))
            if ri % 2 == 0:
                cell.fill = alt_fill
        ws.row_dimensions[ri].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"audit_log_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTS — helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _paginate_list(lst, page, per_page=25):
    """Slice a plain Python list into a pagination dict."""
    total  = len(lst)
    pages  = max(1, (total + per_page - 1) // per_page)
    page   = max(1, min(page, pages))
    start  = (page - 1) * per_page
    return {
        "rows":     lst[start : start + per_page],
        "total":    total,
        "page":     page,
        "pages":    pages,
        "per_page": per_page,
        "has_prev": page > 1,
        "has_next": page < pages,
        "prev_num": page - 1,
        "next_num": page + 1,
    }
 
 
PHT_OFFSET = timedelta(hours=8)
 
def _parse_date_range():
    """Return (date_from, date_to, date_from_str, date_to_str).
    Dates are treated as PHT then converted to UTC for DB queries,
    so 'today' always means the full PHT calendar day.
    """
    today_pht = (datetime.utcnow() + PHT_OFFSET).date()
 
    dfrom_s = request.args.get("date_from", today_pht.isoformat())
    dto_s   = request.args.get("date_to",   today_pht.isoformat())
 
    try:
        dfrom_pht = datetime.strptime(dfrom_s, "%Y-%m-%d").replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        dfrom = dfrom_pht - PHT_OFFSET
    except ValueError:
        dfrom_pht = (datetime.utcnow() + PHT_OFFSET).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        dfrom   = dfrom_pht - PHT_OFFSET
        dfrom_s = dfrom_pht.date().isoformat()
 
    try:
        dto_pht = datetime.strptime(dto_s, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, microsecond=999999
        )
        dto = dto_pht - PHT_OFFSET
    except ValueError:
        dto_pht = (datetime.utcnow() + PHT_OFFSET).replace(
            hour=23, minute=59, second=59, microsecond=999999
        )
        dto   = dto_pht - PHT_OFFSET
        dto_s = dto_pht.date().isoformat()
 
    return dfrom, dto, dfrom_s, dto_s
 


def _build_chart_data(data):
    """
    Produce JSON-serialisable dicts consumed by Chart.js in reports.html.
 
    Key naming contract (must match the JS in reports.html):
        D.sales.labels / revenue / transactions / cost / profit   ← daily arrays
        D.sales.hourly_labels / hourly_revenue / hourly_txn /     ← hourly arrays
                hourly_cost / hourly_profit                         (flat, not nested)
        D.top_products.labels / units
        D.inventory.labels / in_stock / defective / threshold
        D.inventory.category_labels / category_units              ← units by category
        D.inventory.status_ok / status_low / status_out           ← individual counts
        D.stock.labels / units / product_labels / product_units
        D.defects.origin_labels / origin_counts
        D.defects.reason_labels / reason_counts
        D.defects.supplier_labels / supplier_counts
    """
 
    # ── Sales ─────────────────────────────────────────────────────────────────
    daily_asc = list(reversed(data["sales"]["daily"]))  # oldest → newest
    hourly    = data["sales"]["hourly"]                  # [] for multi-day

    sales_chart = {
        # Daily arrays (multi-day view)
        "labels":       [r["date"]        for r in daily_asc],
        "revenue":      [r["total"]        for r in daily_asc],
        "transactions": [r["transactions"] for r in daily_asc],
        "cost":         [r["cost"]         for r in daily_asc],
        "profit":       [r["profit"]       for r in daily_asc],
        # Flat hourly arrays (single-day view) — named so JS can read them as
        # D.sales.hourly_labels etc. without a nested object
        "hourly_labels":  [f"{r['hour']:02d}:00" for r in hourly],
        "hourly_revenue": [r["total"]             for r in hourly],
        "hourly_txn":     [r["transactions"]      for r in hourly],
        "hourly_cost":    [r["cost"]              for r in hourly],
        "hourly_profit":  [r["profit"]            for r in hourly],
    }
 
    top_products_chart = {
        "labels":  [r["product_name"][:22].capitalize() for r in data["sales"]["top_products"]],
        "units":   [int(r["units_sold"]   or 0)          for r in data["sales"]["top_products"]],
        "revenue": [float(r["revenue"]    or 0)          for r in data["sales"]["top_products"]],
    }
 
    # ── Inventory ─────────────────────────────────────────────────────────────
    # Top 12 products by available stock
    inv_sorted = sorted(
        data["inventory"]["rows"],
        key=lambda x: x[1].quantity_available,
        reverse=True,
    )[:12]
 
    inv_chart = {
        "labels":    [p.product_name[:22].capitalize() for p, inv, _ in inv_sorted],
        "in_stock":  [inv.quantity_available            for _, inv, _ in inv_sorted],
        "defective": [(inv.quantity_defective or 0)     for _, inv, _ in inv_sorted],
        "threshold": [p.low_reorder_threshold           for p, _, _ in inv_sorted],
    }
 
    # Category breakdown — units by category (used by D.inventory.category_*)
    cat_units: Counter = Counter()
    for p, inv, cat in data["inventory"]["rows"]:
        label = cat.category_name.capitalize() if cat else "Uncategorised"
        cat_units[label] += inv.quantity_available
    inv_chart["category_labels"] = [k for k, _ in cat_units.most_common()]
    inv_chart["category_units"]  = [v for _, v in cat_units.most_common()]
 
    # Status counts (individual keys so JS can read D.inventory.status_ok etc.)
    all_inv = data["inventory"]["rows"]
    inv_chart["status_ok"]  = sum(
        1 for p, inv, _ in all_inv
        if inv.quantity_available > p.low_reorder_threshold
    )
    inv_chart["status_low"] = data["inventory"]["low_stock"]
    inv_chart["status_out"] = data["inventory"]["out_of_stock"]
 
    # ── Stock In ──────────────────────────────────────────────────────────────
    stock_by_date: dict[str, int] = defaultdict(int)
    for s, _, _ in data["stock"]["rows"]:
        if s.stockin_datetime:
            d = str(to_pht(s.stockin_datetime).date())
            stock_by_date[d] += s.quantity_received
    stock_dates = sorted(stock_by_date)
 
    stock_chart = {
        "labels": stock_dates,
        "units":  [stock_by_date[d] for d in stock_dates],
        "product_labels": [
            r.product_name[:22].capitalize() for r in data["stock"]["by_product"]
        ],
        "product_units":  [
            int(r.units or 0) for r in data["stock"]["by_product"]
        ],
    }
 
    # ── Defects ───────────────────────────────────────────────────────────────
    reason_counter: Counter = Counter()
    for detail, _, _, _ in data["defects"]["rows"]:
        reason_counter[detail.reason.replace("_", " ").title()] += detail.quantity
 
    # Supplier compensation — canonical order matches DB enum values
    _SUP_MAP = {
        "pending":        "Pending",
        "loss":           "Loss",
        "same_item":      "Same Item",
        "different_item": "Diff Item",
        "money":          "Money",
        "none":           "None",
    }
    sup_breakdown = data["defects"].get("supplier_comp_breakdown", {})
    defects_chart = {
        "origin_labels": ["Customer", "In-Store"],
        "origin_counts": [
            data["defects"]["customer_origin"],
            data["defects"]["store_origin"],
        ],
        "reason_labels": [k for k, _ in reason_counter.most_common(8)],
        "reason_counts": [v for _, v in reason_counter.most_common(8)],
        "supplier_labels": list(_SUP_MAP.values()),
        "supplier_counts": [sup_breakdown.get(k, 0) for k in _SUP_MAP],
    }
 
    return {
        "sales":        sales_chart,
        "top_products": top_products_chart,
        "inventory":    inv_chart,
        "stock":        stock_chart,
        "defects":      defects_chart,
    }
 
def _build_report_data(date_from, date_to):
    """Collect all four report-tab datasets for the given date range."""
    from app.models.sale        import Sale
    from app.models.sale_detail import SaleDetail
    from app.models.category    import Category
    from sqlalchemy import func as _f

    is_single_day = (
        (date_from + PHT_OFFSET).date() == (date_to + PHT_OFFSET).date()
    )

    # ── 1. SALES ──────────────────────────────────────────────────────────────
    sales_list = (
        Sale.query
        .filter(Sale.sale_datetime.between(date_from, date_to))
        .order_by(Sale.sale_datetime.desc())
        .all()
    )

    # Daily breakdown (oldest → newest after reversal in _build_chart_data)
    daily_sales = [
        {
            "date":         str(r.date),
            "transactions": int(r.transactions),
            "total":        float(r.total  or 0),
            "cost":         float(r.cost   or 0),
            "profit":       float(r.profit or 0),
        }
        for r in (
            db.session.query(
                _f.date(Sale.sale_datetime).label("date"),
                _f.count(Sale.transaction_id).label("transactions"),
                _f.coalesce(_f.sum(Sale.total_amount),        0).label("total"),
                _f.coalesce(_f.sum(Sale.total_cost_price),    0).label("cost"),
                _f.coalesce(_f.sum(Sale.total_revenue_price), 0).label("profit"),
            )
            .filter(Sale.sale_datetime.between(date_from, date_to))
            .group_by(_f.date(Sale.sale_datetime))
            .order_by(_f.date(Sale.sale_datetime).desc())
            .all()
        )
    ]

    total_revenue = sum(float(s.total_amount        or 0) for s in sales_list)
    total_cost    = sum(float(s.total_cost_price    or 0) for s in sales_list)
    total_profit  = sum(float(s.total_revenue_price or 0) for s in sales_list)
    avg_order     = (total_revenue / len(sales_list)) if sales_list else 0.0

    # Top 10 products by units sold
    gross_units = (
        db.session.query(
            Product.product_id,
            Product.product_name,
            _f.sum(SaleDetail.quantity).label("units_sold"),
            _f.sum(SaleDetail.subtotal_amount).label("revenue"),
        )
        .join(SaleDetail, SaleDetail.product_id     == Product.product_id)
        .join(Sale,       Sale.transaction_id        == SaleDetail.transaction_id)
        .filter(Sale.sale_datetime.between(date_from, date_to))
        .group_by(Product.product_id, Product.product_name)
        .all()
    )

    # ── FIX: anchor returned_units to the ORIGINAL sale_datetime, not defect_datetime
    # This ensures a return processed today deducts from the period the sale occurred,
    # not from whichever period the return was logged in.
    returned_units = (
        db.session.query(
            DefectDetail.product_id,
            _f.sum(DefectDetail.quantity).label("units_returned"),
        )
        .join(Defect,      Defect.defect_id           == DefectDetail.defect_id)
        .join(SaleDetail,  SaleDetail.product_id      == DefectDetail.product_id)
        .join(Sale,        Sale.transaction_id         == SaleDetail.transaction_id)
        .filter(
            Sale.sale_datetime.between(date_from, date_to),   # ← anchored to sale date
            DefectDetail.status      == "active",
            DefectDetail.is_archived == False,
            DefectDetail.origin      == "customer",
            DefectDetail.customer_compensation.in_(
                ["full_refund", "exchange_same", "exchange_different"]
            ),
        )
        .group_by(DefectDetail.product_id)
        .all()
    )

    returned_map = {r.product_id: int(r.units_returned or 0) for r in returned_units}

    top_products_net = []
    for row in gross_units:
        net_units   = max(0, int(row.units_sold or 0) - returned_map.get(row.product_id, 0))
        net_revenue = max(0.0, float(row.revenue or 0))
        top_products_net.append({
            "product_id":   row.product_id,
            "product_name": row.product_name,
            "units_sold":   net_units,
            "revenue":      net_revenue,
        })

    top_products = sorted(top_products_net, key=lambda x: x["units_sold"], reverse=True)[:10]

    total_units_sold_gross = sum(int(r.units_sold or 0) for r in gross_units)
    total_units_returned   = sum(returned_map.values())
    net_units_sold         = max(0, total_units_sold_gross - total_units_returned)

    # ── FIX: no longer deducting defect losses from sales figures.
    # Defect losses live exclusively in defects_data, filtered by defect_datetime,
    # which is correct for the Defects tab. Mixing them here caused cross-period
    # distortion (a return logged today silently reduced a different period's profit).
    # net_* = gross_* intentionally — the Defects tab is the source of truth for losses.

    # Hourly breakdown (single-day only)
    if is_single_day:
        _pht_hour = _f.mod(_f.hour(Sale.sale_datetime) + 8, 24)
        hourly_sales = [
            {
                "hour": int(r.hour) if r.hour is not None else 0,
                "transactions": int(r.transactions),
                "total":        float(r.total  or 0),
                "cost":         float(r.cost   or 0),
                "profit":       float(r.profit or 0),
            }
            for r in (
                db.session.query(
                    _pht_hour.label("hour"),
                    _f.count(Sale.transaction_id).label("transactions"),
                    _f.coalesce(_f.sum(Sale.total_amount),        0).label("total"),
                    _f.coalesce(_f.sum(Sale.total_cost_price),    0).label("cost"),
                    _f.coalesce(_f.sum(Sale.total_revenue_price), 0).label("profit"),
                )
                .filter(Sale.sale_datetime.between(date_from, date_to))
                .group_by(_pht_hour)
                .order_by(_pht_hour)
                .all()
            )
        ]
    else:
        hourly_sales = []

    sales_data = {
        # Gross figures — source of truth for the Sales tab
        "total_revenue":      total_revenue,
        "total_cost":         total_cost,
        "total_profit":       total_profit,
        # net_* kept for frontend compatibility — equal to gross intentionally.
        # Defect losses are reported separately in defects_data, not deducted here.
        "net_revenue":        total_revenue,
        "net_cost":           total_cost,
        "net_profit":         total_profit,
        # Zeroed out — losses belong to the Defects tab, not Sales
        "defect_sales_loss":  0.0,
        "defect_cost_loss":   0.0,
        "defect_profit_loss": 0.0,
        "total_transactions":   len(sales_list),
        "avg_order":            avg_order,
        "total_units_sold":     total_units_sold_gross,
        "net_units_sold":       net_units_sold,
        "total_units_returned": total_units_returned,
        "daily":                daily_sales,
        "hourly":               hourly_sales,
        "top_products":         top_products,
    }

    # ── 2. INVENTORY ──────────────────────────────────────────────────────────
    inv_rows = (
        db.session.query(Product, Inventory, Category)
        .join(Inventory, Inventory.product_id == Product.product_id)
        .outerjoin(Category, Category.category_id == Product.category_id)
        .filter(Product.status != "archived")
        .order_by(Product.product_name)
        .all()
    )

    inventory_data = {
        "total_products":  len(inv_rows),
        "total_units":     sum(inv.quantity_available        for _, inv, _ in inv_rows),
        "total_defective": sum((inv.quantity_defective or 0) for _, inv, _ in inv_rows),
        "low_stock": sum(
            1 for p, inv, _ in inv_rows
            if 0 < inv.quantity_available <= p.low_reorder_threshold
        ),
        "out_of_stock": sum(
            1 for _, inv, _ in inv_rows if inv.quantity_available == 0
        ),
        "rows": inv_rows,
    }

    # ── 3. STOCK-IN ───────────────────────────────────────────────────────────
    stock_rows = (
        db.session.query(StockIn, Product, User)
        .join(Product, Product.product_id == StockIn.product_id)
        .join(User,    User.user_id        == StockIn.user_id)
        .filter(StockIn.stockin_datetime.between(date_from, date_to))
        .order_by(StockIn.stockin_datetime.desc())
        .all()
    )

    stock_by_product = (
        db.session.query(
            Product.product_name,
            _f.sum(StockIn.quantity_received).label("units"),
        )
        .join(Product, Product.product_id == StockIn.product_id)
        .filter(StockIn.stockin_datetime.between(date_from, date_to))
        .group_by(Product.product_id, Product.product_name)
        .order_by(_f.sum(StockIn.quantity_received).desc())
        .limit(8)
        .all()
    )

    stock_data = {
        "total_entries": len({s.batch_id for s, _, _ in stock_rows}),
        "total_units":   sum(s.quantity_received for s, _, _ in stock_rows),
        "by_product":    stock_by_product,
        "rows":          stock_rows,
    }

    # ── 4. DEFECTS ────────────────────────────────────────────────────────────
    # All defect figures are correctly anchored to defect_datetime — no change needed.
    defect_rows = (
        db.session.query(DefectDetail, Defect, Product, User)
        .join(Defect,  Defect.defect_id   == DefectDetail.defect_id)
        .join(Product, Product.product_id == DefectDetail.product_id)
        .join(User,    User.user_id        == Defect.user_id)
        .filter(Defect.defect_datetime.between(date_from, date_to))
        .filter(DefectDetail.is_archived == False)
        .order_by(Defect.defect_datetime.desc())
        .all()
    )

    supplier_comp: Counter = Counter()
    for d, _, _, _ in defect_rows:
        if d.status == "active":
            supplier_comp[d.supplier_compensation] += 1

    active_defect_rows = [(d, df, p, u) for d, df, p, u in defect_rows if d.status == "active"]

    defects_data = {
        "total":           len(defect_rows),
        "total_units":     sum(d.quantity for d, _, _, _ in defect_rows),
        "customer_origin": sum(1 for d, _, _, _ in defect_rows if d.origin == "customer"),
        "store_origin":    sum(1 for d, _, _, _ in defect_rows if d.origin == "in_store"),
        "total_sales_loss": sum(
            float(d.subtotal_amount or 0)
            if d.customer_compensation == "full_refund"
            else abs(float(d.price_difference or 0))
            for d, _, _, _ in active_defect_rows
            if d.origin == "customer"
            and d.customer_compensation in ("full_refund", "partial_refund")
        ),
        "supplier_recovered": sum(
            float(d.subtotal_unit or 0)
            for d, _, _, _ in active_defect_rows
            if d.supplier_compensation in ("money", "exchange_same", "exchange_different")
        ),
        "supplier_pending": sum(
            float(d.subtotal_unit or 0)
            for d, _, _, _ in active_defect_rows
            if d.supplier_compensation == "pending"
        ),
        "total_cost_loss": sum(
            float(d.subtotal_unit or 0) for d, _, _, _ in active_defect_rows
        ),
        "total_profit_loss": sum(
            float(d.subtotal_revenue or 0)
            if d.customer_compensation == "full_refund"
            else abs(float(d.price_difference or 0))
            for d, _, _, _ in active_defect_rows
            if d.origin == "customer"
            and d.customer_compensation in ("full_refund", "partial_refund")
        ),
        "supplier_loss": sum(
            float(d.subtotal_unit or 0)
            for d, _, _, _ in active_defect_rows
            if d.supplier_compensation == "loss"
        ),
        "supplier_comp_breakdown": dict(supplier_comp),
        "rows": defect_rows,
    }

    return {
        "sales":     sales_data,
        "inventory": inventory_data,
        "stock":     stock_data,
        "defects":   defects_data,
    }
 
 
@admin_bp.route("/reports/data")
@login_required
@role_required("superadmin", "admin")
def reports_data():
    """
    JSON feed consumed by the reports dashboard frontend.
 
    Query params (all optional – fall back to today):
        date_from   YYYY-MM-DD
        date_to     YYYY-MM-DD
 
    Returns a single JSON object with four top-level keys:
        date_from, date_to, sales, inventory, stock, defects
    """
    date_from, date_to, date_from_str, date_to_str = _parse_date_range()
    data = _build_report_data(date_from, date_to)
 
    s   = data["sales"]
    inv = data["inventory"]
    sk  = data["stock"]
    df  = data["defects"]
 
    # ── Sales rows (latest 50 transactions) ───────────────────────────────────
    # Uses Sale model imported inside _build_report_data; re-import cleanly here.
    from app.models.sale import Sale
 
    sales_rows_json = [
        {
            "transaction_id":    sale.transaction_id,
            # PHT-localised display string
            "sale_datetime":     to_pht(sale.sale_datetime).strftime("%b %d, %Y %I:%M %p")
                                 if sale.sale_datetime else "—",
            "user":              f"{sale.user.first_name} {sale.user.last_name}".strip().title()
                                 if sale.user else "—",
            # ── KEY COLUMNS ──────────────────────────────────────────────────
            # total_amount       = what the customer paid          (Revenue)
            # total_cost_price   = cost of goods sold              (Cost)
            # total_revenue_price = markup / profit component      (Profit)
            # Invariant: total_amount = total_cost_price + total_revenue_price
            "total_amount":         float(sale.total_amount        or 0),
            "total_cost_price":     float(sale.total_cost_price    or 0),
            "total_revenue_price":  float(sale.total_revenue_price or 0),
            "payment_method":       sale.payment_method or "—",
        }
        for sale in (
            Sale.query
            .filter(Sale.sale_datetime.between(date_from, date_to))
            .order_by(Sale.sale_datetime.desc())
            .limit(50)
            .all()
        )
    ]
 
    # ── Top products (already queried in _build_report_data) ─────────────────
    top_products_json = [
        {
            "product_name": r["product_name"],
            "units_sold":   r["units_sold"],    # already net (returns subtracted)
            "revenue":      r["revenue"],
        }
        for r in s["top_products"]
    ]
 
 
    # ── Inventory rows ────────────────────────────────────────────────────────
    inv_rows_json = [
        {
            "product_id":             product.product_id,
            "product_name":           product.product_name,
            "category_name":          category.category_name if category else "—",
            # Current stock counts (not date-filtered – always current state)
            "quantity_available":     inventory.quantity_available,
            "quantity_defective":     inventory.quantity_defective or 0,
            "low_reorder_threshold":  product.low_reorder_threshold,
            "status":                 product.status,
            "last_updated":           to_pht(inventory.last_updated).strftime("%b %d, %Y")
                                      if inventory.last_updated else "—",
        }
        for product, inventory, category in inv["rows"]
    ]
 
    # ── Stock In rows ─────────────────────────────────────────────────────────
    stock_rows_json = [
        {
            "stockin_id":        stock_in.stockin_id,          # PK: stockin_id (NOT stock_in_id)
            "stockin_datetime":  to_pht(stock_in.stockin_datetime).strftime("%b %d, %Y %I:%M %p")
                                 if stock_in.stockin_datetime else "—",
            "product_name":      product.product_name,
            "quantity_received": stock_in.quantity_received,   # field: quantity_received
            "user":              f"{user.first_name} {user.last_name}".strip().title(),
            "notes":             stock_in.notes or "—",
            "batch_id":          stock_in.batch_id,            # nullable FK → Stock_In_Batch
        }
        for stock_in, product, user in sk["rows"]
    ]
 
    # ── Stock by-product breakdown (top 8 for chart) ─────────────────────────
    stock_by_product_json = [
        {
            "product_name": r.product_name,
            "units":        int(r.units or 0),
        }
        for r in sk["by_product"]
    ]
 
    # ── Defect rows ───────────────────────────────────────────────────────────
    # Financial columns:
    #   subtotal_amount   = selling-price loss  (used for "sales loss")
    #   subtotal_unit     = cost loss            (used for "cost loss")
    #   subtotal_revenue  = profit/markup loss   (used for "profit loss")
    #
    # IMPORTANT: aggregate loss figures are for status='active' records ONLY.
    # submitted / rejected records are EXCLUDED from all ₱ loss totals.
    defect_rows_json = [
        {
            "defect_detail_id":       detail.defect_detail_id,
            "defect_datetime":        to_pht(defect.defect_datetime).strftime("%b %d, %Y %I:%M %p")
                                      if defect.defect_datetime else "—",
            "product_name":           product.product_name,
            "quantity":               detail.quantity,
            "origin":                 detail.origin,                # 'customer' | 'in_store'
            "reason":                 detail.reason,                # 'damaged' | 'expired' | 'change_of_mind'
            "status":                 detail.status,                # 'submitted' | 'active' | 'rejected'
            "subtotal_amount":        float(detail.subtotal_amount  or 0),
            "subtotal_unit":          float(detail.subtotal_unit    or 0),
            "subtotal_revenue":       float(detail.subtotal_revenue or 0),
            "customer_compensation":  detail.customer_compensation,
            "supplier_compensation":  detail.supplier_compensation,
            "user":                   f"{user.first_name} {user.last_name}".strip().title(),
        }
        for detail, defect, product, user in df["rows"]
    ]
 
    return jsonify({
        "date_from": date_from_str,
        "date_to":   date_to_str,
 
        # ── Sales ──────────────────────────────────────────────────────────
        "sales": {
            # Aggregate totals
            "total_revenue":       float(s["total_revenue"]),      # sum(total_amount)
            "total_cost":          float(s["total_cost"]),         # sum(total_cost_price)
            "total_profit":        float(s["total_profit"]),       # sum(total_revenue_price)
            "total_transactions":   s["total_transactions"],
            "avg_order":            float(s["avg_order"]),
            "total_units_sold":     s["total_units_sold"],    # gross
            "net_units_sold":       s["net_units_sold"],      # after returns ← use this on frontend
            "total_units_returned": s["total_units_returned"],
            # Defect deductions (active defects only)
            "defect_sales_loss":   float(s["defect_sales_loss"]),  # sum(subtotal_amount)  active
            "defect_cost_loss":    float(s["defect_cost_loss"]),   # sum(subtotal_unit)    active
            "defect_profit_loss":  float(s["defect_profit_loss"]), # sum(subtotal_revenue) active
            # Chart series
            "daily":          s["daily"],     # [{date, transactions, total, cost, profit}, ...]
            "hourly":         s["hourly"],    # populated only when date_from == date_to
            "top_products":   top_products_json,
            "rows":           sales_rows_json,
        },
 
        # ── Inventory (current state, not date-filtered) ────────────────
        "inventory": {
            "total_products": inv["total_products"],
            "total_units":    inv["total_units"],    # sum(quantity_available)
            "low_stock":      inv["low_stock"],      # 0 < qty_available <= low_reorder_threshold
            "out_of_stock":   inv["out_of_stock"],   # qty_available == 0
            "rows":           inv_rows_json,
        },
 
        # ── Stock In ───────────────────────────────────────────────────────
        "stock": {
            "total_entries": sk["total_entries"],
            "total_units":   sk["total_units"],      # sum(quantity_received)
            "by_product":    stock_by_product_json,
            "rows":          stock_rows_json,
        },
 
        # ── Defects ────────────────────────────────────────────────────────
        "defects": {
            "total":              df["total"],
            "total_units":        df["total_units"],
            "customer_origin":    df["customer_origin"],
            "store_origin":       df["store_origin"],
            # Loss totals — status='active' records only
            "total_sales_loss":   float(df["total_sales_loss"]),   # sum(subtotal_amount)
            "total_cost_loss":    float(df["total_cost_loss"]),    # sum(subtotal_unit)
            "total_profit_loss":  float(df["total_profit_loss"]),  # sum(subtotal_revenue)
            "supplier_loss":      float(df["supplier_loss"]),      # where supplier_compensation='loss'
            "rows":               defect_rows_json,
        },
    })


@admin_bp.route("/audit_logs/<int:log_id>")
@login_required
@role_required("superadmin", "admin")
def audit_log_detail(log_id):
    """Return a single audit log entry as JSON for the detail modal."""
    log = AuditLog.query.get_or_404(log_id)
    u   = log.user

    # Fetch email from Recovery_Details
    email = "—"
    if u and u.recovery_detail:
        email = u.recovery_detail.email or "—"

    return jsonify({
        "log_id":          log.log_id,
        "action_datetime": to_pht(log.action_datetime).strftime("%B %d, %Y  %I:%M:%S %p PHT")
                           if log.action_datetime else "—",
        "action_type":     log.action_type or "—",
        "module":          log.module or "—",
        "description":     log.description or "—",
        "reference_table": log.reference_table or None,
        "reference_id":    log.reference_id or None,
        "user": {
            "name":  f"{u.first_name} {u.last_name}".strip().title() if u else "—",
            "role":  u.role.title() if u else "—",
            "email": email,
            "id":    u.user_id if u else None,
        },
    })


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTS — routes
# ═══════════════════════════════════════════════════════════════════════════════

@admin_bp.route("/reports/pdf")
@login_required
@role_required("superadmin", "admin")
def export_report_pdf():
    """Render a print-ready PDF for the current tab + date range."""
    from weasyprint import HTML

    date_from, date_to, date_from_str, date_to_str = _parse_date_range()
    tab  = request.args.get("tab", "sales")
    data = _build_report_data(date_from, date_to)

    html_str = render_template(
        "admin/reports_pdf.html",
        data         = data,
        tab          = tab,
        date_from    = date_from_str,
        date_to      = date_to_str,
        generated_at = to_pht(datetime.utcnow()).strftime("%B %d, %Y  %I:%M %p PHT"),
    )

    pdf_bytes = HTML(string=html_str, base_url=request.host_url).write_pdf()

    buf = io.BytesIO(pdf_bytes)
    buf.seek(0)
    filename = f"report_{tab}_{date_from_str}_to_{date_to_str}.pdf"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
    )

@admin_bp.route("/reports/export")
@login_required
@role_required("superadmin", "admin")
def export_report():
    """Download the selected report tab as a polished, color-coded Excel workbook."""
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

    date_from, date_to, date_from_str, date_to_str = _parse_date_range()
    tab  = request.args.get("tab", "sales")
    data = _build_report_data(date_from, date_to)

    wb = openpyxl.Workbook()

    # ── Palette ───────────────────────────────────────────────────────────────
    CLR = {
        "header_bg":   "1E293B",   # dark slate header
        "header_fg":   "FFFFFF",
        "accent":      "6366F1",   # indigo accent
        "row_alt":     "F8FAFC",
        "border":      "E2E8F0",
        "title_bg":    "F1F5F9",
        "green_bg":    "D1FAE5",   "green_fg":    "065F46",
        "amber_bg":    "FEF3C7",   "amber_fg":    "92400E",
        "red_bg":      "FEE2E2",   "red_fg":      "991B1B",
        "blue_bg":     "DBEAFE",   "blue_fg":     "1E40AF",
        "orange_bg":   "FFEDD5",   "orange_fg":   "9A3412",
        "violet_bg":   "EDE9FE",   "violet_fg":   "5B21B6",
        "rose_bg":     "FFE4E6",   "rose_fg":     "9F1239",
        "green_num":   "059669",
        "red_num":     "DC2626",
        "orange_num":  "EA580C",
        "amber_num":   "D97706",
    }

    CURRENCY_FMT = u'₱#,##0.00'
    INT_FMT      = '#,##0'

    # ── Style helpers ─────────────────────────────────────────────────────────
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border(color="E2E8F0", style="thin"):
        s = Side(style=style, color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def thick_bottom(color="CBD5E1"):
        b = Side(style="medium", color=color)
        return Border(bottom=b)

    def font(bold=False, size=10, color="1E293B", italic=False):
        return Font(bold=bold, size=size, color=color, italic=italic,
                    name="Calibri")

    def align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def write_title_block(ws, title, subtitle, row=1):
        """Write a branded title block at the top of a sheet."""
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font      = Font(bold=True, size=16, color=CLR["header_bg"], name="Calibri")
        c.fill      = fill(CLR["title_bg"])
        c.alignment = align("left", "center")
        ws.row_dimensions[row].height = 30

        ws.merge_cells(f"A{row+1}:H{row+1}")
        c2 = ws.cell(row=row+1, column=1, value=subtitle)
        c2.font      = Font(italic=True, size=9, color="64748B", name="Calibri")
        c2.fill      = fill(CLR["title_bg"])
        c2.alignment = align("left", "center")
        ws.row_dimensions[row+1].height = 16
        return row + 3   # return next available row (leaves one blank)

    def write_kv_block(ws, pairs, start_row):
        """Write key-value summary rows with styled labels."""
        for i, (label, value, fmt) in enumerate(pairs):
            r = start_row + i
            lc = ws.cell(row=r, column=1, value=label)
            lc.font      = font(bold=True, size=9, color="64748B")
            lc.fill      = fill("F8FAFC")
            lc.alignment = align("left")
            lc.border    = border()

            vc = ws.cell(row=r, column=2, value=value)
            vc.font      = font(bold=True, size=11)
            vc.fill      = fill("FFFFFF")
            vc.alignment = align("right")
            vc.border    = border()
            if fmt:
                vc.number_format = fmt
            ws.row_dimensions[r].height = 18
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 20
        return start_row + len(pairs) + 1

    def make_table_header(ws, row, headers):
        """Write a styled table header row. headers = [(label, width, align)]"""
        for ci, (h, w, ha) in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font      = Font(bold=True, color=CLR["header_fg"], size=9, name="Calibri")
            c.fill      = fill(CLR["header_bg"])
            c.alignment = Alignment(horizontal=ha, vertical="center")
            c.border    = border(CLR["header_bg"])
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[row].height = 22
        ws.freeze_panes = ws.cell(row=row+1, column=1)

    def write_data_row(ws, row_idx, values, formats=None, colors=None, height=17):
        """Write one data row. values = list, formats = list of num fmts or None,
           colors = list of (bg_hex, fg_hex) or None per cell."""
        alt = row_idx % 2 == 0
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.border    = border()
            c.alignment = Alignment(vertical="center")

            # alternating row fill (only if no specific color override)
            if colors and ci <= len(colors) and colors[ci-1]:
                bg, fg = colors[ci-1]
                c.fill = fill(bg)
                c.font = Font(color=fg, size=9, name="Calibri", bold=True)
            else:
                c.fill = fill(CLR["row_alt"]) if alt else fill("FFFFFF")
                c.font = Font(color="334155", size=9, name="Calibri")

            if formats and ci <= len(formats) and formats[ci-1]:
                c.number_format = formats[ci-1]
        ws.row_dimensions[row_idx].height = height

    def write_totals_row(ws, row_idx, values, formats=None):
        """Write a bold totals / subtotals footer row."""
        for ci, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.fill      = fill("E2E8F0")
            c.font      = Font(bold=True, size=9, color=CLR["header_bg"], name="Calibri")
            c.border    = Border(top=Side(style="medium", color="94A3B8"),
                                  bottom=Side(style="medium", color="94A3B8"))
            c.alignment = Alignment(horizontal="right" if ci > 1 else "left",
                                     vertical="center")
            if formats and ci <= len(formats) and formats[ci-1]:
                c.number_format = formats[ci-1]
        ws.row_dimensions[row_idx].height = 20

    def badge_fill(status_str):
        """Return (bg, fg) color pair for a status string."""
        s = status_str.lower()
        if s in ("active", "ok", "approved"):
            return CLR["green_bg"], CLR["green_fg"]
        if s in ("rejected", "out"):
            return CLR["red_bg"], CLR["red_fg"]
        if s in ("low",):
            return CLR["amber_bg"], CLR["amber_fg"]
        if s in ("submitted", "pending", "partially_approved"):
            return CLR["blue_bg"], CLR["blue_fg"]
        return "F1F5F9", "334155"

    # ═══════════════════════════════════════════════════════════════════════════
    # SALES
    # ═══════════════════════════════════════════════════════════════════════════
    if tab == "sales":
        s = data["sales"]

        # ── Sheet 1: Summary ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        ws.sheet_view.showGridLines = False

        nxt = write_title_block(ws, "Sales Report", f"Period: {date_from_str}  →  {date_to_str}")

        kv_pairs = [
            ("Total Transactions",     s["total_transactions"],           INT_FMT),
            ("Total Revenue (₱)",      float(s["total_revenue"]),         CURRENCY_FMT),
            ("Total Cost (₱)",         float(s.get("total_cost", 0)),     CURRENCY_FMT),
            ("Gross Profit (₱)",       float(s.get("total_profit", 0)),   CURRENCY_FMT),
            ("Avg Order Value (₱)",    float(s["avg_order"]),             CURRENCY_FMT),
            ("Total Units Sold",       s.get("total_units_sold", 0),      INT_FMT),
            ("Defect: Sales Loss (₱)", float(s.get("defect_sales_loss",  0)), CURRENCY_FMT),
            ("Defect: Cost Loss (₱)",  float(s.get("defect_cost_loss",   0)), CURRENCY_FMT),
            ("Defect: Profit Loss (₱)",float(s.get("defect_profit_loss", 0)), CURRENCY_FMT),
        ]
        nxt = write_kv_block(ws, kv_pairs, nxt)

        # Profit margin calculation in a separate highlighted cell
        rev   = float(s["total_revenue"])
        prof  = float(s.get("total_profit", 0))
        margin = (prof / rev * 100) if rev > 0 else 0
        margin_row = nxt
        ws.cell(row=margin_row, column=1, value="Gross Margin (%)").font = font(bold=True, size=9, color="64748B")
        ws.cell(row=margin_row, column=1).fill = fill("F8FAFC")
        mc = ws.cell(row=margin_row, column=2, value=round(margin, 2))
        mc.font = Font(bold=True, size=14, color=CLR["green_num"] if margin >= 0 else CLR["red_num"], name="Calibri")
        mc.number_format = "0.00\"%\""
        mc.fill = fill("ECFDF5" if margin >= 0 else "FEF2F2")
        mc.alignment = align("right")
        ws.row_dimensions[margin_row].height = 26

        # ── Sheet 2: Daily Breakdown ───────────────────────────────────────────
        ws2 = wb.create_sheet("Daily Breakdown")
        ws2.sheet_view.showGridLines = False

        write_title_block(ws2, "Daily Sales Breakdown", f"Period: {date_from_str}  →  {date_to_str}")

        hdrs = [
            ("Date",              16, "left"),
            ("Transactions",      15, "right"),
            ("Revenue (₱)",       18, "right"),
            ("Cost (₱)",          16, "right"),
            ("Profit (₱)",        16, "right"),
            ("Margin (%)",        13, "right"),
        ]
        make_table_header(ws2, 4, hdrs)

        tot_txn = tot_rev = tot_cost = tot_profit = 0
        for ri, row in enumerate(s["daily"], 5):
            rev_r  = float(row.get("total",  0))
            cost_r = float(row.get("cost",   0))
            prof_r = float(row.get("profit", 0))
            mgn    = (prof_r / rev_r * 100) if rev_r > 0 else 0
            tot_txn    += row["transactions"]
            tot_rev    += rev_r
            tot_cost   += cost_r
            tot_profit += prof_r
            write_data_row(ws2, ri,
                [row["date"], row["transactions"], rev_r, cost_r, prof_r, round(mgn, 2)],
                formats=[None, INT_FMT, CURRENCY_FMT, CURRENCY_FMT, CURRENCY_FMT, '0.00"%"'],
                colors=[None, None, None,
                        ("FFEDD5", CLR["orange_num"]),
                        ("ECFDF5" if prof_r >= 0 else "FEF2F2",
                         CLR["green_num"] if prof_r >= 0 else CLR["red_num"]),
                        None])
        tot_margin = (tot_profit / tot_rev * 100) if tot_rev > 0 else 0
        write_totals_row(ws2, len(s["daily"]) + 5,
            ["TOTAL", tot_txn, tot_rev, tot_cost, tot_profit, round(tot_margin, 2)],
            formats=[None, INT_FMT, CURRENCY_FMT, CURRENCY_FMT, CURRENCY_FMT, '0.00"%"'])

        # ── Sheet 3: Top Products ──────────────────────────────────────────────
        ws3 = wb.create_sheet("Top Products")
        ws3.sheet_view.showGridLines = False
        write_title_block(ws3, "Top Products by Revenue", f"Period: {date_from_str}  →  {date_to_str}")

        make_table_header(ws3, 4, [
            ("Rank",        8,  "center"),
            ("Product",     38, "left"),
            ("Units Sold",  14, "right"),
            ("Revenue (₱)", 18, "right"),
        ])
        for ri, row in enumerate(s["top_products"], 5):
            rev_val = float(row.get("revenue") or 0)
            write_data_row(ws3, ri,
                [ri - 4, row["product_name"].capitalize(),
                 int(row.get("units_sold") or 0), rev_val],
                formats=[None, None, INT_FMT, CURRENCY_FMT])

    # ═══════════════════════════════════════════════════════════════════════════
    # INVENTORY
    # ═══════════════════════════════════════════════════════════════════════════
    elif tab == "inventory":
        inv = data["inventory"]

        ws = wb.active
        ws.title = "Inventory"
        ws.sheet_view.showGridLines = False
        nxt = write_title_block(ws, "Inventory Report", f"Snapshot as of: {date_to_str}")

        write_kv_block(ws, [
            ("Total Products",  inv["total_products"], INT_FMT),
            ("Total Units",     inv["total_units"],    INT_FMT),
            ("Low Stock Items", inv["low_stock"],      INT_FMT),
            ("Out of Stock",    inv["out_of_stock"],   INT_FMT),
        ], nxt)

        # ── Sheet 2: Full product list ─────────────────────────────────────────
        ws2 = wb.create_sheet("All Products")
        ws2.sheet_view.showGridLines = False
        write_title_block(ws2, "Product Inventory Detail", f"Snapshot as of: {date_to_str}")

        hdrs = [
            ("Product",            36, "left"),
            ("Category",           20, "left"),
            ("In Stock",           12, "right"),
            ("Defective",          12, "right"),
            ("Reorder Threshold",  18, "right"),
            ("Status",             12, "center"),
            ("Product Status",     14, "center"),
            ("Last Updated",       20, "left"),
        ]
        make_table_header(ws2, 4, hdrs)

        for ri, (product, inv_row, category) in enumerate(inv["rows"], 5):
            qty   = inv_row.quantity_available
            thresh= product.low_reorder_threshold
            is_out = qty == 0
            is_low = not is_out and qty <= thresh

            stock_status = "Out" if is_out else ("Low" if is_low else "OK")
            sbg, sfg     = badge_fill(stock_status)
            last_upd     = to_pht(inv_row.last_updated).strftime("%b %d, %Y") if inv_row.last_updated else "—"

            # qty color based on stock level
            qty_color = None
            if is_out:
                qty_color = (CLR["red_bg"], CLR["red_fg"])
            elif is_low:
                qty_color = (CLR["amber_bg"], CLR["amber_fg"])
            else:
                qty_color = ("ECFDF5", CLR["green_num"])

            write_data_row(ws2, ri, [
                product.product_name.capitalize(),
                category.category_name.capitalize() if category else "—",
                qty,
                inv_row.quantity_defective or 0,
                thresh,
                stock_status,
                product.status.title(),
                last_upd,
            ],
            formats=[None, None, INT_FMT, INT_FMT, INT_FMT, None, None, None],
            colors=[None, None, qty_color, None, None,
                    (sbg, sfg),
                    badge_fill(product.status),
                    None])

        # ── Sheet 3: Action Required ───────────────────────────────────────────
        ws3 = wb.create_sheet("⚠ Reorder Required")
        ws3.sheet_view.showGridLines = False
        write_title_block(ws3, "Reorder Required", f"Out-of-stock & low-stock items — {date_to_str}")

        make_table_header(ws3, 4, [
            ("Product",             36, "left"),
            ("Category",            20, "left"),
            ("In Stock",            12, "right"),
            ("Reorder Threshold",   18, "right"),
            ("Units Needed",        14, "right"),
            ("Status",              12, "center"),
        ])
        ri = 5
        for product, inv_row, category in inv["rows"]:
            qty   = inv_row.quantity_available
            thresh = product.low_reorder_threshold
            is_out = qty == 0
            is_low = not is_out and qty <= thresh
            if not is_out and not is_low:
                continue
            needed = thresh - qty
            stock_status = "Out" if is_out else "Low"
            sbg, sfg = badge_fill(stock_status)
            write_data_row(ws3, ri, [
                product.product_name.capitalize(),
                category.category_name.capitalize() if category else "—",
                qty,
                thresh,
                needed,
                stock_status,
            ],
            formats=[None, None, INT_FMT, INT_FMT, INT_FMT, None],
            colors=[None, None,
                    (CLR["red_bg"], CLR["red_fg"]) if is_out else (CLR["amber_bg"], CLR["amber_fg"]),
                    None,
                    (CLR["amber_bg"], CLR["amber_fg"]),
                    (sbg, sfg)])
            ri += 1

    # ═══════════════════════════════════════════════════════════════════════════
    # STOCK IN
    # ═══════════════════════════════════════════════════════════════════════════
    elif tab == "stock":
        sk = data["stock"]

        ws = wb.active
        ws.title = "Stock Movement"
        ws.sheet_view.showGridLines = False
        nxt = write_title_block(ws, "Stock-In Report", f"Period: {date_from_str}  →  {date_to_str}")

        write_kv_block(ws, [
            ("Period",              f"{date_from_str}  →  {date_to_str}", None),
            ("Total Entries",       sk["total_entries"],                   INT_FMT),
            ("Total Units Received",sk["total_units"],                     INT_FMT),
        ], nxt)

        # ── Sheet 2: Records ───────────────────────────────────────────────────
        ws2 = wb.create_sheet("Records")
        ws2.sheet_view.showGridLines = False
        write_title_block(ws2, "Stock-In Records", f"Period: {date_from_str}  →  {date_to_str}")

        hdrs = [
            ("Date & Time (PHT)",  22, "left"),
            ("Product",            36, "left"),
            ("Qty Received",       14, "right"),
            ("Received By",        22, "left"),
            ("Batch ID",           10, "center"),
            ("Notes",              40, "left"),
        ]
        make_table_header(ws2, 4, hdrs)

        total_qty = 0
        for ri, (stock_in, product, user) in enumerate(sk["rows"], 5):
            pht_dt = (to_pht(stock_in.stockin_datetime).strftime("%b %d, %Y  %I:%M %p")
                      if stock_in.stockin_datetime else "—")
            name   = f"{user.first_name} {user.last_name}".strip().title()
            total_qty += stock_in.quantity_received
            write_data_row(ws2, ri, [
                pht_dt,
                product.product_name.capitalize(),
                stock_in.quantity_received,
                name,
                stock_in.batch_id or "—",
                stock_in.notes or "—",
            ],
            formats=[None, None, INT_FMT, None, None, None],
            colors=[None, None, ("ECFDF5", CLR["green_num"]), None, None, None],
            height=18)
        write_totals_row(ws2, len(sk["rows"]) + 5,
            ["TOTAL", "", total_qty, "", "", ""],
            formats=[None, None, INT_FMT, None, None, None])

    # ═══════════════════════════════════════════════════════════════════════════
    # DEFECTS
    # ═══════════════════════════════════════════════════════════════════════════
    elif tab == "defects":
        df = data["defects"]

        # ── Sheet 1: Summary ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        ws.sheet_view.showGridLines = False
        nxt = write_title_block(ws, "Defects Report", f"Period: {date_from_str}  →  {date_to_str}")

        write_kv_block(ws, [
            ("Period",                  f"{date_from_str}  →  {date_to_str}", None),
            ("Total Defect Records",    df["total"],                           INT_FMT),
            ("Total Units Defective",   df["total_units"],                     INT_FMT),
            ("Customer Origin",         df["customer_origin"],                 INT_FMT),
            ("In-Store Origin",         df["store_origin"],                    INT_FMT),
            ("",                        "",                                    None),
            ("Sales Loss (₱)",          float(df.get("total_sales_loss",  0)), CURRENCY_FMT),
            ("Profit Loss (₱)",         float(df.get("total_profit_loss", 0)), CURRENCY_FMT),
            ("Cost Loss (₱)",           float(df.get("total_cost_loss",   0)), CURRENCY_FMT),
            ("",                        "",                                    None),
            ("Supplier Recovered (₱)",  float(df.get("supplier_recovered",0)), CURRENCY_FMT),
            ("Supplier Absorbed (₱)",   float(df.get("supplier_loss",     0)), CURRENCY_FMT),
            ("Supplier Pending (₱)",    float(df.get("supplier_pending",  0)), CURRENCY_FMT),
        ], nxt)

        # ── Sheet 2: Records ───────────────────────────────────────────────────
        ws2 = wb.create_sheet("Records")
        ws2.sheet_view.showGridLines = False
        write_title_block(ws2, "Defect Records", f"Period: {date_from_str}  →  {date_to_str}")

        hdrs = [
            ("Date (PHT)",          20, "left"),
            ("Product",             30, "left"),
            ("Qty",                  8, "right"),
            ("Origin",              12, "center"),
            ("Reason",              20, "left"),
            ("Sales Loss (₱)",      16, "right"),
            ("Cost Loss (₱)",       16, "right"),
            ("Profit Loss (₱)",     16, "right"),
            ("Cust. Compensation",  22, "left"),
            ("Supp. Compensation",  22, "left"),
            ("Status",              12, "center"),
            ("Logged By",           22, "left"),
        ]
        make_table_header(ws2, 4, hdrs)

        tot_qty = tot_sales = tot_cost = tot_profit = 0
        for ri, (detail, defect, product, user) in enumerate(df["rows"], 5):
            pht_dt  = (to_pht(defect.defect_datetime).strftime("%b %d, %Y")
                       if defect.defect_datetime else "—")
            name    = f"{user.first_name} {user.last_name}".strip().title()
            sales_v = float(detail.subtotal_amount  or 0)
            cost_v  = float(detail.subtotal_unit    or 0)
            prof_v  = float(detail.subtotal_revenue or 0)
            tot_qty    += detail.quantity
            tot_sales  += sales_v
            tot_cost   += cost_v
            tot_profit += prof_v

            origin_color = (CLR["rose_bg"],   CLR["rose_fg"])   if detail.origin == "customer" \
                      else (CLR["orange_bg"], CLR["orange_fg"])
            sbg, sfg = badge_fill(detail.status)

            write_data_row(ws2, ri, [
                pht_dt,
                product.product_name.capitalize(),
                detail.quantity,
                "Customer" if detail.origin == "customer" else "In-Store",
                detail.reason.replace("_", " ").title(),
                sales_v,
                cost_v,
                prof_v,
                detail.customer_compensation.replace("_", " ").title(),
                detail.supplier_compensation.replace("_", " ").title(),
                detail.status.title(),
                name,
            ],
            formats=[None, None, INT_FMT, None, None,
                     CURRENCY_FMT, CURRENCY_FMT, CURRENCY_FMT,
                     None, None, None, None],
            colors=[None, None, None,
                    origin_color,
                    None,
                    (CLR["red_bg"],   CLR["red_fg"]),
                    ("FFEDD5",        CLR["orange_num"]),
                    (CLR["red_bg"],   CLR["red_fg"]),
                    None, None,
                    (sbg, sfg),
                    None])

        write_totals_row(ws2, len(df["rows"]) + 5,
            ["TOTAL", "", tot_qty, "", "", tot_sales, tot_cost, tot_profit, "", "", "", ""],
            formats=[None, None, INT_FMT, None, None,
                     CURRENCY_FMT, CURRENCY_FMT, CURRENCY_FMT,
                     None, None, None, None])

        # ── Sheet 3: By Product ────────────────────────────────────────────────
        # Aggregate defect totals per product from rows
        from collections import defaultdict
        by_product = defaultdict(lambda: {"qty": 0, "sales": 0.0, "cost": 0.0, "profit": 0.0})
        for detail, defect, product, user in df["rows"]:
            key = product.product_name.capitalize()
            by_product[key]["qty"]    += detail.quantity
            by_product[key]["sales"]  += float(detail.subtotal_amount  or 0)
            by_product[key]["cost"]   += float(detail.subtotal_unit    or 0)
            by_product[key]["profit"] += float(detail.subtotal_revenue or 0)
        sorted_products = sorted(by_product.items(), key=lambda x: x[1]["qty"], reverse=True)

        ws3 = wb.create_sheet("By Product")
        ws3.sheet_view.showGridLines = False
        write_title_block(ws3, "Defects by Product", f"Period: {date_from_str}  →  {date_to_str}")

        make_table_header(ws3, 4, [
            ("Product",         36, "left"),
            ("Total Qty",       12, "right"),
            ("Sales Loss (₱)",  16, "right"),
            ("Cost Loss (₱)",   16, "right"),
            ("Profit Loss (₱)", 16, "right"),
        ])
        for ri, (pname, vals) in enumerate(sorted_products, 5):
            write_data_row(ws3, ri,
                [pname, vals["qty"], vals["sales"], vals["cost"], vals["profit"]],
                formats=[None, INT_FMT, CURRENCY_FMT, CURRENCY_FMT, CURRENCY_FMT],
                colors=[None, None,
                        (CLR["red_bg"], CLR["red_fg"]),
                        ("FFEDD5", CLR["orange_num"]),
                        (CLR["red_bg"], CLR["red_fg"])])

    else:
        wb.active.title = "No Data"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"report_{tab}_{date_from_str}_to_{date_to_str}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
# ═══════════════════════════════════════════════════════════════════════════════
# SALES HISTORY  (admin / superadmin)
# ═══════════════════════════════════════════════════════════════════════════════

@admin_bp.route("/sales")
@login_required
@role_required("superadmin", "admin")
def sales_history():
    from app.models.sale        import Sale
    from app.models.sale_detail import SaleDetail
    from sqlalchemy             import func
    from datetime               import date as _date, timedelta

    PER_PAGE = 50

    page      = request.args.get("page",      1,  type=int)
    q         = request.args.get("q",         "").strip()
    user_id   = request.args.get("user_id",   "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to   = request.args.get("date_to",   "").strip()

    qry = (
        Sale.query
        .join(User, Sale.user_id == User.user_id)
        .order_by(Sale.sale_datetime.desc())
    )

    if user_id:
        try:
            qry = qry.filter(Sale.user_id == int(user_id))
        except ValueError:
            user_id = ""

    if q:
        try:
            qry = qry.filter(Sale.transaction_id == int(q))
        except ValueError:
            pass

    if date_from:
        try:
            qry = qry.filter(
                Sale.sale_datetime >= datetime.strptime(date_from, "%Y-%m-%d")
            )
        except ValueError:
            date_from = ""

    if date_to:
        try:
            qry = qry.filter(
                Sale.sale_datetime
                <= datetime.strptime(date_to, "%Y-%m-%d")
                           .replace(hour=23, minute=59, second=59)
            )
        except ValueError:
            date_to = ""

    sales = qry.paginate(page=page, per_page=PER_PAGE, error_out=False)

    for sale in sales.items:
        sale.has_override = any(d.override_used for d in sale.sale_details)

    # ── Period stats (no user filter — whole store) ───────────────────────
    today_dt    = datetime.combine(_date.today(), datetime.min.time())
    week_dt     = today_dt - timedelta(days=today_dt.weekday())
    month_dt    = today_dt.replace(day=1)
    year_dt     = today_dt.replace(month=1, day=1)

    def _period(start):
        row = db.session.query(
            func.count(Sale.transaction_id),
            func.coalesce(func.sum(Sale.total_amount),        0),
            func.coalesce(func.sum(Sale.total_cost_price),    0),
            func.coalesce(func.sum(Sale.total_revenue_price), 0),
        ).filter(Sale.sale_datetime >= start).first()
        return dict(
            count  = int(row[0]),
            sales  = float(row[1]),
            cost   = float(row[2]),
            profit = float(row[3]),
        )

    stats = dict(
        today = _period(today_dt),
        week  = _period(week_dt),
        month = _period(month_dt),
        year  = _period(year_dt),
    )

    all_users = (
        User.query
        .filter(User.role.in_(["superadmin", "admin", "cashier"]))
        .order_by(User.first_name)
        .all()
    )

    filters = dict(
        q         = q,
        user_id   = int(user_id) if user_id and user_id.isdigit() else None,
        date_from = date_from,
        date_to   = date_to,
    )

    return render_template(
        "sales/history.html",
        sales     = sales,
        filters   = filters,
        stats     = stats,
        all_users = all_users,
        is_admin  = True,
    )


@admin_bp.route("/sales/<int:transaction_id>")
@login_required
@role_required("superadmin", "admin")
def sale_detail(transaction_id):
    from app.models.sale        import Sale
    from app.models.sale_detail import SaleDetail

    sale = Sale.query.get_or_404(transaction_id)
    u    = sale.user

    details = [
    {
        "product_id":   d.product_id,
        "product_name": d.product.product_name.capitalize() if d.product else d.product_id,
        "quantity":     d.quantity,
        "price":        float(d.price_at_sale),
        "subtotal":     float(d.subtotal_amount),
        "override_used": d.override_used,  
    }
    for d in sale.sale_details
]

    return jsonify({
        "transaction_id":      sale.transaction_id,
        "sale_datetime":       to_pht(sale.sale_datetime).strftime("%B %d, %Y  %I:%M:%S %p PHT")
                               if sale.sale_datetime else "—",
        "total_amount":        float(sale.total_amount),
        "total_cost_price":    float(sale.total_cost_price),
        "total_revenue_price": float(sale.total_revenue_price),
        "tendered_amount":     float(sale.tendered_amount) if sale.tendered_amount is not None else None,
        "change_amount":       float(sale.change_amount)   if sale.change_amount   is not None else None,
        "payment_method":      sale.payment_method or "cash",
        "cashier": {
            "name": f"{u.first_name} {u.last_name}".strip().title() if u else "—",
            "role": u.role.title() if u else "—",
            "id":   u.user_id if u else None,
        },
        "items": details,
    })

@admin_bp.get("/notifications/count")
@login_required
def notifications_count():
    since_ts = request.args.get("since", type=float)
    entries, count = get_navbar_notifications(
        current_user.role, since_ts=since_ts
    )
    return {"count": count}


@admin_bp.route('/products/bulk-import', methods=['GET', 'POST'])
@login_required
@role_required("superadmin")
def bulk_import_products():
    if request.method == 'GET':
        return render_template('settings/bulk_import.html')

    file = request.files.get('csv_file')
    if not file or not file.filename.endswith('.csv'):
        flash('Please upload a valid .csv file.', 'danger')
        return redirect(request.url)

    stream = io.StringIO(file.stream.read().decode('utf-8-sig'))
    reader = csv.DictReader(stream)
    reader.fieldnames = [h.strip().lower() for h in reader.fieldnames]

    REQUIRED = {'barcode', 'name', 'category', 'cost price', 'markup price'}
    if not REQUIRED.issubset(set(reader.fieldnames)):
        flash(f'CSV is missing required columns. Expected: {", ".join(REQUIRED)}', 'danger')
        return redirect(request.url)

    categories   = {c.category_name.strip().lower(): c for c in Category.query.all()}
    existing_ids = {p.product_id for p in db.session.query(Product.product_id).all()}
    existing_names = {p.product_name.strip().lower() for p in db.session.query(Product.product_name).all()}
    seen_names_in_csv = set()

    rows_created = 0
    rows_updated = 0
    error_rows   = []

    for raw in reader:
        row = {k.strip(): (v.strip() if v else '') for k, v in raw.items()}

        barcode      = row.get('barcode', '').strip()
        name         = row.get('name', '').strip()
        category_raw = row.get('category', '').strip()
        cost_raw     = row.get('cost price', '').strip()
        markup_raw   = row.get('markup price', '').strip()

        reason = None

        if not barcode:
            reason = 'missing barcode'
        elif not name:
            reason = 'missing name'
        elif name.lower() in existing_names and barcode not in existing_ids:
            # existing_ids check: allow updates to existing barcodes even if name matches
            reason = 'duplicate name (already exists in database)'
        elif name.lower() in seen_names_in_csv:
            reason = 'duplicate name (appears more than once in this file)'

        else:
            try:
                cost_price   = round(float(cost_raw)   if cost_raw   else 0.0, 2)
                markup_price = round(float(markup_raw) if markup_raw else 0.0, 2)
                if cost_price < 0 or markup_price < 0:
                    raise ValueError
            except ValueError:
                reason = 'invalid cost or markup price'

            if reason is None:
                cat = categories.get(category_raw.lower())
                if cat is None and category_raw:
                    new_cat = Category(
                        category_name               = category_raw,
                        description                 = None,
                        status                      = 'active',
                        default_low_stock_threshold = 5,
                    )
                    db.session.add(new_cat)
                    db.session.flush()
                    categories[category_raw.lower()] = new_cat
                    cat = new_cat
                elif not category_raw:
                    reason = 'missing category'

        if reason:
            error_rows.append({
                'barcode':      barcode,
                'name':         name,
                'category':     category_raw,
                'cost price':   cost_raw,
                'markup price': markup_raw,
                'reason':       reason,
            })
            continue

        total_price = round(cost_price + markup_price, 2)

        if barcode in existing_ids:
            product = Product.query.get(barcode)
            if product:
                product.cost_price    = cost_price
                product.revenue_price = markup_price
                product.total_price   = total_price
                rows_updated += 1
        else:
            product = Product(
                product_id            = barcode,
                product_name          = name,
                category_id           = cat.category_id,
                cost_price            = cost_price,
                revenue_price         = markup_price,
                total_price           = total_price,
                low_reorder_threshold = cat.default_low_stock_threshold,
                status                = 'active',
                created_at            = datetime.utcnow(),
            )
            inventory = Inventory(
                product_id         = barcode,
                quantity_available = 0,
                quantity_defective = 0,
                last_updated       = datetime.utcnow(),
            )
            db.session.add(product)
            db.session.add(inventory)
            existing_ids.add(barcode)
            existing_names.add(name.lower())   # ADD THIS
            seen_names_in_csv.add(name.lower()) # ADD THIS
            rows_created += 1

    # ── commit once after all rows are processed ─────────────────────────────
    db.session.commit()

    if error_rows:
        out = io.StringIO()
        writer = csv.DictWriter(
            out,
            fieldnames=['barcode', 'name', 'category', 'cost price', 'markup price', 'reason']
        )
        writer.writeheader()
        writer.writerows(error_rows)

        flash(
            f'{rows_created} product(s) created, {rows_updated} updated. '
            f'{len(error_rows)} row(s) had errors — see downloaded file.',
            'warning'
        )

        return Response(
            out.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': 'attachment; filename="import_errors.csv"',
                'X-Import-Success': str(rows_created + rows_updated),
                'X-Import-Errors':  str(len(error_rows)),
            }
        )

    flash(
        f'{rows_created} product(s) created, {rows_updated} updated successfully.',
        'success'
    )
    return redirect(url_for('admin.bulk_import_products'))


# ── NO indentation — this must be at module level ────────────────────────────
@admin_bp.route('/products/export', methods=['GET'])
@login_required
@role_required("superadmin")
def export_products():
    products = (
        db.session.query(Product, Category)
        .outerjoin(Category, Product.category_id == Category.category_id)
        .filter(Product.status == 'active')
        .order_by(Product.product_name)
        .all()
    )

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(['barcode', 'name', 'category', 'cost price', 'markup price'])

    for product, category in products:
        writer.writerow([
            product.product_id,
            product.product_name,
            category.category_name if category else '',
            product.cost_price,
            product.revenue_price,
        ])

    return Response(
        out.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment; filename="products_export.csv"'
        }
    )